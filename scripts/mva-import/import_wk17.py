#!/usr/bin/env python3
"""
Import week-17 (4/19–4/25/2026) Pedestrian Struck + Traffic Stops from xlsx.

Sources:
  /Users/geremy/Desktop/IMPACT April 2026/Week 4:19 - 4:25/PedStruckWeek17.xlsx
  /Users/geremy/Desktop/IMPACT April 2026/Week 4:19 - 4:25/Traffic Stop week 17 .xlsx

PII STRIPPING — none of these source columns are written to data.json:
  • Person Name (Type)        — includes JUVENILE flags
  • Assigned Name / Officer   — primary + secondary officer assignments
  • Case Disposition / notes  — may contain narrative

Stored per record:
  Pedestrian Struck → id (ps3-XXXX), type, date, district, lat, lng,
                      address (block-level), description (HH:MM time)
  Traffic Stop      → id (ts3-XXXX), type, date, district, lat, lng,
                      address (block-level, no city suffix)

Usage:
  python3 scripts/mva-import/import_wk17.py            # dry run
  python3 scripts/mva-import/import_wk17.py --apply
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path
from typing import Optional

import openpyxl

ROOT       = Path(__file__).resolve().parents[2]
DATA_JSON  = ROOT / 'src' / 'data' / 'data.json'
CACHE_FILE = ROOT / 'scripts' / 'geocode-cache.json'

# Source files are copied to /tmp by the operator (Python sandbox can't always
# reach the user's Desktop directly).
# Originals: ~/Desktop/IMPACT April 2026/Week 4:19 - 4:25/{PedStruckWeek17.xlsx, Traffic Stop week 17 .xlsx}
PED_XLSX = Path('/tmp/ped_wk17.xlsx')
TS_XLSX  = Path('/tmp/ts_wk17.xlsx')

ARCGIS_URL = 'https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates'
JC_BOX     = (40.66, 40.77, -74.13, -74.02)
MIN_SCORE  = 75

# Shared helpers
sys.path.insert(0, str(ROOT / 'scripts' / 'mva-import'))
sys.path.insert(0, str(ROOT / 'scripts' / 'privacy'))
from _stats import regenerate_stats  # noqa: E402
from block_address import to_block    # noqa: E402


# ── Geocoding ─────────────────────────────────────────────────────────────

def in_jc(lat, lng):
    return JC_BOX[0] <= lat <= JC_BOX[1] and JC_BOX[2] <= lng <= JC_BOX[3]


def arcgis_geocode(address: str):
    params = urllib.parse.urlencode({
        'SingleLine': address, 'outFields': 'Score,Addr_type', 'maxLocations': 3,
        'forStorage': 'false', 'f': 'json', 'location': '-74.0776,40.7282',
        'distance': 20000, 'countryCode': 'USA',
    })
    try:
        req = urllib.request.Request(f'{ARCGIS_URL}?{params}', headers={'User-Agent': 'JCImpact/2.0'})
        with urllib.request.urlopen(req, timeout=12) as r:
            data = json.loads(r.read())
        for c in data.get('candidates', []):
            score = c.get('score', 0)
            loc = c.get('location') or {}
            lat, lng = loc.get('y'), loc.get('x')
            if score >= MIN_SCORE and lat and lng and in_jc(lat, lng):
                return lat, lng
    except Exception as e:
        print(f'    ArcGIS error: {e}')
    return None, None


# ── District PIP ──────────────────────────────────────────────────────────

_DISTRICTS_CACHE = None
def load_districts():
    global _DISTRICTS_CACHE
    if _DISTRICTS_CACHE is not None: return _DISTRICTS_CACHE
    src = (ROOT / 'src' / 'data' / 'boundaryData.ts').read_text()
    m = re.search(r'districtGeoJSON\s*(?::\s*[^=]+)?\s*=\s*', src)
    start = m.end(); depth = 0; i = start
    while i < len(src):
        ch = src[i]
        if ch == '{': depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0: i += 1; break
        i += 1
    _DISTRICTS_CACHE = json.loads(src[start:i])['features']
    return _DISTRICTS_CACHE

def pip_ring(pt, ring):
    x, y = pt; inside = False; n = len(ring); j = n - 1
    for k in range(n):
        xi, yi = ring[k]; xj, yj = ring[j]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi):
            inside = not inside
        j = k
    return inside

def pip(pt, coords, t):
    if t == 'Polygon':
        if not pip_ring(pt, coords[0]): return False
        for h in coords[1:]:
            if pip_ring(pt, h): return False
        return True
    if t == 'MultiPolygon':
        return any(pip(pt, p, 'Polygon') for p in coords)
    return False

def district_for(lng, lat) -> Optional[str]:
    for f in load_districts():
        if pip((lng, lat), f['geometry']['coordinates'], f['geometry']['type']):
            return f['properties']['name']
    return None


# ── xlsx parsing helpers ──────────────────────────────────────────────────

def parse_dt(v) -> Optional[dt.datetime]:
    if isinstance(v, dt.datetime): return v
    if isinstance(v, str):
        for fmt in ('%m/%d/%Y %H:%M', '%m/%d/%Y'):
            try: return dt.datetime.strptime(v.strip(), fmt)
            except Exception: pass
    return None


def clean_ped_location(raw: str) -> str:
    """Normalize the messy "Location&Address / Loc. Type" cell.

    Examples:
      "\nGRAND ST & CLINTON AVE, JERSEY CITY, NJ 07304-0000 (STREET)"
      "COLUMBUS DR and\nCOLUMBUS DR, JERSEY CITY, NJ 07302 (STREET)"
      "\n18 FULTON AVE, JERSEY CITY, NJ 07305 (STREET) at Garfield"
    """
    s = raw.strip()
    # Replace "and" between newlines with " & "
    s = re.sub(r'\s+and\s+', ' & ', s, flags=re.IGNORECASE)
    # Collapse newlines and excess whitespace
    s = re.sub(r'\s*\n\s*', ' ', s)
    # Strip trailing "(STREET)" or similar location-type marker, and any "at X" suffix
    s = re.sub(r'\s*\([A-Z]+\)\s*(at\s+.*)?$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    # Collapse "X & X" duplications (same street twice — geocoder confusion)
    parts = [p.strip() for p in s.split('&')]
    if len(parts) > 1 and parts[0].split(',')[0].strip().upper() == parts[1].split(',')[0].strip().upper():
        # Same street on both sides — keep just one half (with city suffix)
        s = parts[1] if ',' in parts[1] else parts[0]
    return s


def strip_city_suffix(raw: str) -> str:
    """For traffic stops — match the existing TS address format (no city/state/zip)."""
    s = re.sub(r',?\s*JERSEY CITY.*$', '', raw, flags=re.IGNORECASE).strip()
    s = re.sub(r',?\s*NJ\s*\d{5}.*$', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r',\s*NJ\s*$', '', s, flags=re.IGNORECASE).strip()
    return s


# ── Main ─────────────────────────────────────────────────────────────────

def run(apply: bool):
    data = json.loads(DATA_JSON.read_text())
    try:
        cache = json.loads(CACHE_FILE.read_text())
    except Exception:
        cache = {}

    # Next IDs
    max_ps3 = max((int(re.match(r'ps3-(\d+)$', i['id']).group(1))
                   for i in data['incidents']
                   if re.match(r'ps3-(\d+)$', i.get('id', ''))), default=0)
    max_ts3 = max((int(re.match(r'ts3-(\d+)$', i['id']).group(1))
                   for i in data['incidents']
                   if re.match(r'ts3-(\d+)$', i.get('id', ''))), default=0)

    # Track existing case_numbers for ped struck dedupe (ps3 doesn't store case#,
    # so we fall back to (date, address) match)
    existing_ps_keys = {(i['date'], i['address'].upper())
                        for i in data['incidents'] if i['type'] == 'Pedestrian Struck'}

    new_incidents: list[dict] = []
    failed: list[dict] = []
    api_calls = 0
    cache_hits = 0

    def geocode(query: str):
        nonlocal api_calls, cache_hits
        key = query.lower().strip()
        hit = cache.get(key)
        if hit:
            cache_hits += 1
            return hit['lat'], hit['lng']
        if not apply:
            return None, None  # skip API in dry run
        lat, lng = arcgis_geocode(query)
        api_calls += 1
        time.sleep(0.25)
        if lat and lng:
            cache[key] = {'lat': lat, 'lng': lng, 'cached_at': '2026-04-28'}
        return lat, lng

    # ── PEDESTRIAN STRUCK ──
    print('=== Pedestrian Struck ===')
    wb = openpyxl.load_workbook(PED_XLSX, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True)
            if any(v is not None and str(v).strip() for v in r)]
    header, ped_rows = rows[0], rows[1:]
    print(f'Source rows: {len(ped_rows)}')

    next_ps = max_ps3 + 1
    for r in ped_rows:
        # Columns: 0=Report Date/Time, 1=Case#, 2=Person Name (PII), 3=Officer (PII),
        #          4=Location&Address / Loc. Type, 5=District
        when = parse_dt(r[0])
        if not when:
            failed.append({'kind': 'ped', 'reason': 'unparseable date', 'raw': r[0]})
            continue
        loc_raw = (r[4] or '').strip()
        if not loc_raw:
            failed.append({'kind': 'ped', 'reason': 'empty location', 'case': r[1]})
            continue
        loc = clean_ped_location(loc_raw)

        date_iso = when.date().isoformat()
        time_str = when.strftime('%H:%M')

        # Dedupe
        stored_addr_for_key = to_block(loc).upper()
        if (date_iso, stored_addr_for_key) in existing_ps_keys:
            failed.append({'kind': 'ped', 'reason': 'already imported', 'case': r[1]})
            continue

        lat, lng = geocode(loc)
        if not (lat and lng):
            failed.append({'kind': 'ped', 'reason': 'geocode failed', 'query': loc, 'case': r[1]})
            continue

        dist = district_for(lng, lat)
        if not dist:
            xlsx_dist = (r[5] or '').strip().title()
            if xlsx_dist in ('North', 'East', 'South', 'West'):
                dist = xlsx_dist
            else:
                failed.append({'kind': 'ped', 'reason': 'outside districts', 'case': r[1]})
                continue

        # PII STRIP: do NOT include person names, officer names, or juvenile flags
        new_incidents.append({
            'id': f'ps3-{next_ps:04d}',
            'type': 'Pedestrian Struck',
            'date': date_iso,
            'district': dist,
            'lat': lat,
            'lng': lng,
            'address': to_block(loc),
            'description': time_str,
        })
        existing_ps_keys.add((date_iso, stored_addr_for_key))
        next_ps += 1

    # ── TRAFFIC STOPS ──
    print('\n=== Traffic Stops ===')
    wb = openpyxl.load_workbook(TS_XLSX, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True)
            if any(v is not None and str(v).strip() for v in r)]
    header, ts_rows = rows[0], rows[1:]
    print(f'Source rows: {len(ts_rows)}')

    next_ts = max_ts3 + 1
    for r in ts_rows:
        # Columns: 0=Report?, 1=Report Date/Time, 2=Incident Location,
        #          3=CFS Type, 4=Disposition
        when = parse_dt(r[1])
        if not when:
            failed.append({'kind': 'ts', 'reason': 'unparseable date', 'raw': r[1]})
            continue
        raw_loc = (r[2] or '').strip()
        if not raw_loc:
            failed.append({'kind': 'ts', 'reason': 'empty location'})
            continue

        date_iso = when.date().isoformat()
        # Storage form: strip city suffix THEN apply block transform
        stored_addr = to_block(strip_city_suffix(raw_loc))

        lat, lng = geocode(raw_loc)
        if not (lat and lng):
            failed.append({'kind': 'ts', 'reason': 'geocode failed', 'query': raw_loc})
            continue

        dist = district_for(lng, lat)
        if not dist:
            failed.append({'kind': 'ts', 'reason': 'outside districts', 'query': raw_loc})
            continue

        new_incidents.append({
            'id': f'ts3-{next_ts:04d}',
            'type': 'Traffic Stop',
            'date': date_iso,
            'district': dist,
            'lat': lat,
            'lng': lng,
            'address': stored_addr,
        })
        next_ts += 1

    # ── Report ──
    ped_added = sum(1 for n in new_incidents if n['type'] == 'Pedestrian Struck')
    ts_added  = sum(1 for n in new_incidents if n['type'] == 'Traffic Stop')

    print(f'\n┌─ summary ─────────────────────────────')
    print(f'│ Pedestrian Struck added: {ped_added}')
    print(f'│ Traffic Stops added:     {ts_added}')
    print(f'│ Cache hits:              {cache_hits}')
    print(f'│ API calls:               {api_calls}')
    print(f'│ Failed/skipped:          {len(failed)}')
    print(f'└───────────────────────────────────────')

    for f_ in failed[:15]:
        print(f'  - {f_["kind"]}: {f_["reason"]}  {f_.get("case", f_.get("query", ""))!r}')
    if len(failed) > 15:
        print(f'  ... ({len(failed) - 15} more)')

    if not apply:
        print('\n=== DRY RUN — use --apply to write')
        return

    data['incidents'].extend(new_incidents)
    regenerate_stats(data)
    DATA_JSON.write_text(json.dumps(data))
    CACHE_FILE.write_text(json.dumps(cache, indent=2, sort_keys=True))

    # District distribution
    print(f'\nNew records by district:')
    for kind in ('Pedestrian Struck', 'Traffic Stop'):
        c = Counter(n['district'] for n in new_incidents if n['type'] == kind)
        print(f'  {kind:18s}  {dict(c)}')

    print(f'\nNew totals:')
    print(f'  Pedestrian Struck total: {sum(1 for i in data["incidents"] if i["type"] == "Pedestrian Struck")}')
    print(f'  Traffic Stop total:      {sum(1 for i in data["incidents"] if i["type"] == "Traffic Stop")}')
    print(f'  All incidents:           {len(data["incidents"])}')


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--apply', action='store_true')
    args = p.parse_args()
    run(apply=args.apply)


if __name__ == '__main__':
    sys.exit(main() or 0)
