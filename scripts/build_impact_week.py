#!/usr/bin/env python3
"""
build-impact-week.py — Build a JC IMPACT weekly workbook from raw RMS category exports.

Usage:
  python3 scripts/build-impact-week.py /Users/geremy/Desktop/Week\\ 4:12\\ -\\ 4:18 \
      --shapefile ~/Desktop/JCPD\ District\ Shape\ File/us.nj.jersey_city.geojson \
      --output ~/Desktop/Week16_IMPACT_Workbook.xlsx \
      --week 16

What it does:
  1. Scans the week folder for category Excel files (Arrests, Robbery, Burglary, Theft,
     Gun Recoveries, Shoplifting, Agg Assault, Shots Fired, Crash Report, Stolen Vehicle).
  2. Parses each file with category-specific heuristics.
  3. Geocodes addresses via Nominatim (OpenStreetMap) with caching + rate limiting.
  4. Performs point-in-polygon test against JCPD district boundaries.
  5. Aggregates counts: citywide and per-district (North/East/South/West).
  6. Builds Shots Fired & Shootings incident list.
  7. Builds MVA Data incident list (from Crash Report PDF → CSV if available, else skip).
  8. Writes an Excel workbook matching the "JCIMPACT Weekly Jan Feb 2026.xlsx" schema.
"""

import argparse, json, os, re, sys, time
from collections import defaultdict
from datetime import date
from pathlib import Path

import pandas as pd
import requests

# ─── Configuration ─────────────────────────────────────────────────────────────
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
USER_AGENT = "JCIMPACT-Hermes/1.0 (+http://jerseycity.gov)"

# District centroids fallback (used when geocode fails)
CENTROIDS = {
    'North': (40.7378, -74.0533),
    'East':  (40.7178, -74.0450),
    'South': (40.7040, -74.0680),
    'West':  (40.7220, -74.0870),
}

# The 7 major IMPACT categories (all others are supplemental)
MAJOR_CATEGORIES = ['arrests', 'robbery', 'burglary', 'theft',
                    'agg_assault', 'homicide', 'sex_offenses']
# We also track: gun_recoveries, shoplifting, shots_fired, mva, stolen_vehicle

# ─── Geocoding cache ────────────────────────────────────────────────────────────
def load_cache(cache_path):
    if os.path.exists(cache_path):
        with open(cache_path) as f:
            return json.load(f)
    return {}

def save_cache(cache, cache_path):
    with open(cache_path, 'w') as f:
        json.dump(cache, f, indent=2)

def normalize_address(addr: str) -> str:
    """Clean address for geocoding."""
    import re
    # Remove parentheses content
    addr = re.sub(r'\(.*?\)', '', addr)
    # Remove unit/apt/floor info
    addr = re.sub(r',?\s+FL\s+\d+', '', addr, flags=re.I)
    addr = re.sub(r',?\s+APT\s+\w+', '', addr, flags=re.I)
    addr = re.sub(r',?\s+UNIT\s+\w+', '', addr, flags=re.I)
    addr = re.sub(r',?\s+STE\s+\w+', '', addr, flags=re.I)
    addr = re.sub(r',?\s+SUITE\s+\w+', '', addr, flags=re.I)
    addr = re.sub(r',?\s+ROOM\s+\w+', '', addr, flags=re.I)
    addr = re.sub(r',?\s+BSMT', '', addr, flags=re.I)
    addr = re.sub(r',?\s+BASEMENT', '', addr, flags=re.I)
    addr = re.sub(r'\s+', ' ', addr).strip().strip(',')
    if not re.search(r'JERSEY CITY', addr, re.I):
        addr = addr + ', JERSEY CITY, NJ'
    return addr

def geocode_address(addr, cache, session=None):
    key = addr.lower().strip()
    if key in cache:
        return cache[key]  # (lat, lon) or None

    cleaned = normalize_address(addr)
    params = {'q': cleaned, 'format': 'json', 'limit': 1, 'addressdetails': 1}
    headers = {'User-Agent': USER_AGENT}

    try:
        time.sleep(1.12)  # rate limit 1/sec
        resp = (session or requests).get(NOMINATIM_URL, params=params, headers=headers, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if data:
                coords = (float(data[0]['lat']), float(data[0]['lon']))
                cache[key] = coords
                return coords
    except Exception as e:
        print(f"  Geocode error for '{addr[:40]}': {e}", file=sys.stderr)
    cache[key] = None
    return None

# ─── Point-in-polygon ───────────────────────────────────────────────────────────
def point_in_polygon(x, y, poly):
    """Ray-casting: x=lon, y=lat. poly = list of [lon,lat] vertices."""
    inside = False
    n = len(poly)
    j = n - 1
    for i in range(n):
        xi, yi = poly[i]
        xj, yj = poly[j]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside

def load_districts(geojson_path):
    import json
    with open(geojson_path) as f:
        data = json.load(f)
    districts = {}
    for feat in data['features']:
        full = feat['properties'].get('name', '')
        key = full.replace('District', '').strip().title()
        geom = feat['geometry']
        if geom['type'] == 'Polygon':
            polys = [geom['coordinates'][0]]
        elif geom['type'] == 'MultiPolygon':
            polys = [p[0] for p in geom['coordinates']]
        else:
            polys = []
        districts[key] = polys
    return districts

def assign_district(lat, lon, districts_polys):
    for dist_name, polys in districts_polys.items():
        for poly in polys:
            if point_in_polygon(lon, lat, poly):
                return dist_name
    return None

# ─── Date parsing ───────────────────────────────────────────────────────────────
def parse_date(val):
    """Parse mm/dd or mm/dd/yyyy strings → YYYY-MM-DD (2026 default year)."""
    import re
    s = str(val).strip()
    m = re.match(r'(\d{1,2})/(\d{1,2})', s)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        return f'2026-{month:02d}-{day:02d}'
    # Try mm/dd/yyyy
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = year + 2000 if year < 100 else year
        return f'{year}-{month:02d}-{day:02d}'
    return '2026-01-01'

# ─── Category parsers ───────────────────────────────────────────────────────────
CATEGORY_COLS = {
    # category: {date_col_heuristic, address_col_heuristic, district_col_heuristic, case_col_heuristic}
    'robbery':      (r'date',           r'location',      r'district',    r'case'),
    'burglary':     (r'date',           r'location',      None,           r'case'),
    'theft':        (r'date',           r'location',      None,           r'case'),
    'agg_assault':  (r'date',           r'location',      None,           r'case'),
    'arrests':      (r'date.*time',     r'incident.*location', r'district', r'case'),  # Arrests has Incident Date & Time
    'gun':          (r'date',           r'location',      None,           r'case'),
    'shoplifting':  (r'date',           r'location',      None,           r'case'),
    'stolen_vehicle':(r'date',          r'location',      r'district',    r'case'),
}

def infer_columns(headers):
    """Return dict of col indices for required fields for a given category pattern set."""
    lower_map = {i: str(h).strip().lower() for i, h in enumerate(headers) if h}
    fields = {}
    # date: first col whose string looks like a date header or contains 'date'
    for i, h in lower_map.items():
        if re.search(r'date', h) or re.search(r'\d{1,2}/\d{1,2}', str(h)):
            fields['date'] = i
            break
    # address: contains 'location' or 'address'
    for i, h in lower_map.items():
        if 'location' in h or 'address' in h:
            fields['address'] = i
            break
    # case: contains 'case'
    for i, h in lower_map.items():
        if 'case' in h:
            fields['case'] = i
            break
    # district: contains 'district'
    for i, h in lower_map.items():
        if 'district' in h:
            fields['district'] = i
            break
    return fields

def parse_excel_file(filepath, category):
    """Read Excel and return list of incident dicts with raw strings."""
    import pandas as pd
    # read with header=None to get all rows
    try:
        df = pd.read_excel(filepath, header=None, dtype=str)
    except Exception as e:
        print(f"  ERROR reading {filepath}: {e}", file=sys.stderr)
        return []

    if df.empty:
        return []

    # Try header row = row 0
    headers = [str(h).strip() if pd.notna(h) else '' for h in df.iloc[0]]
    # infer columns
    cols = infer_columns(headers)
    # If no address col found, try row 1 as header?
    if 'address' not in cols and len(df) > 1:
        headers2 = [str(h).strip() if pd.notna(h) else '' for h in df.iloc[1]]
        cols2 = infer_columns(headers2)
        if 'address' in cols2:
            # header row is row 1
            headers = headers2
            start_row = 2
            cols = cols2
        else:
            start_row = 1
    else:
        start_row = 1

    # Ensure required cols exist
    if 'address' not in cols:
        print(f"  WARNING: no address column found in {os.path.basename(filepath)} — skipping", file=sys.stderr)
        return []

    records = []
    for idx in range(start_row, len(df)):
        row = df.iloc[idx]
        # Skip completely empty rows
        if row.isna().all():
            continue
        # Get fields
        date_raw = str(row[cols.get('date', 0)]).strip() if 'date' in cols else ''
        addr_raw = str(row[cols['address']]).strip()
        case_raw = str(row[cols.get('case', 0)]).strip() if 'case' in cols else ''
        district_raw = str(row[cols['district']]).strip() if 'district' in cols else ''

        if not addr_raw or addr_raw.lower() in ('nan', 'none'):
            continue

        record = {
            'date': parse_date(date_raw) if date_raw else '',
            'address': addr_raw,
            'case': case_raw,
            'district_direct': district_raw.title() if district_raw else None,
            'category': category,
        }
        records.append(record)
    return records

# ─── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Build IMPACT weekly workbook from raw RMS category files')
    parser.add_argument('week_folder', help='Folder containing weekly category Excel files')
    parser.add_argument('--shapefile', required=True, help='JCPD District GeoJSON path')
    parser.add_argument('--output', required=True, help='Output .xlsx workbook path')
    parser.add_argument('--week', type=int, required=True, help='Week number (e.g., 16)')
    parser.add_argument('--date-range', help='Week date range, e.g. "Apr 12-18 2026"')
    args = parser.parse_args()

    week_folder = Path(args.week_folder)
    cache_path = f'/tmp/geocache_w{args.week}.json'
    districts_polys = load_districts(args.shapefile)
    geocache = load_cache(cache_path)
    session = requests.Session()

    # ── Step 1: Parse all category files ────────────────────────────────────────
    category_files = {
        'arrests':      find_file(week_folder, r'arrest',  re.I),
        'robbery':      find_file(week_folder, r'robbery', re.I),
        'burglary':     find_file(week_folder, r'burglary',re.I),
        'theft':        find_file(week_folder, r'theft(?!.*shop)', re.I),
        'agg_assault':  find_file(week_folder, r'agg.*assault|aggravated.*assault', re.I),
        'gun':          find_file(week_folder, r'gun|firearm|recoveries', re.I),
        'shoplifting':  find_file(week_folder, r'shoplift', re.I),
        'shots_fired':  find_file(week_folder, r'shot.*fired|shooting', re.I),
        'stolen_vehicle': find_file(week_folder, r'stolen.*vehicle', re.I),
        'mva':          find_file(week_folder, r'crash.*report|mva', re.I),
    }

    all_incidents = []    # For shots + MVA incident lists
    category_counts = defaultdict(lambda: {'citywide': 0, 'districts': {d:0 for d in ['North','East','South','West']}})

    for cat, filepath in category_files.items():
        if not filepath or not Path(filepath).exists():
            print(f"  [{cat.upper()}] not found — skipping")
            continue
        print(f"\nParsing {cat}: {Path(filepath).name}")
        records = parse_excel_file(filepath, category)
        print(f"  {len(records)} raw records")
        if not records:
            continue

        # Assign districts
        for rec in records:
            latlon = None
            if rec.get('district_direct') and rec['district_direct'].title() in districts_polys:
                dist = rec['district_direct'].title()
            else:
                georesult = geocode_address(rec['address'], geocache, session)
                if georesult:
                    lat, lon = georesult
                    dist = assign_district(lat, lon, districts_polys)
                    latlon = (lat, lon)
                else:
                    dist = None
            # Save incident-level for shots / MVA if needed
            if cat in ('shots_fired', 'mva'):
                rec['district'] = dist
                rec['lat'], rec['lng'] = latlon if latlon else (None, None)
                all_incidents.append(rec)
            # Count
            category_counts[cat]['citywide'] += 1
            if dist:
                category_counts[cat]['districts'][dist] += 1
        print(f"  → Citywide: {category_counts[cat]['citywide']}  Districts: {category_counts[cat]['districts']}")

    # Save updated cache
    save_cache(geocache, cache_path)

    # ── Step 2: Build summary totals by month ────────────────────────────────────
    # The week folder likely contains data for a single week (7 days). For the
    # CITYWIDE sheet we still need monthly totals. For now, we treat this week
    # as its month bucket. You can later accumulate multiple weeks.
    # Infer month from first incident date
    month_key = None
    for rec in all_incidents:
        if rec.get('date'):
            month_key = rec['date'][:7]  # "2026-04"
            break
    if not month_key:
        month_key = date.today().strftime('%Y-%m')

    # Build the DataFrame-like structures to write with openpyxl
    # We'll construct the workbook in memory with openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # Remove the auto-created sheet
    default = wb.active
    wb.remove(default)

    # ── Helpers to format sheets ─────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")   # medium blue
    ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")   # light blue
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    BOLD_FONT    = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT   = Font(bold=True, size=14, color="1F4E79")
    CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT         = Alignment(horizontal='left',   vertical='center')
    thin = Side(style='thin', color='BFBFBF')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, title, data_grid):
        """data_grid is list of rows, each row is list of cell values."""
        ws.title = title
        for r, row in enumerate(data_grid, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                # Style rules
                if r == 1:  # title row
                    cell.font = TITLE_FONT
                    cell.alignment = CENTER
                elif r == 2:  # subtitle/prepared row
                    cell.alignment = LEFT
                elif r <= 4:
                    pass
                else:
                    cell.alignment = LEFT
                cell.border = BORDER

    # ── CITYWIDE sheet ───────────────────────────────────────────────────────────
    citywide_grid = build_citywide_grid(category_counts, month_key, week_num=args.week)
    write_sheet(wb.create_sheet("CITYWIDE"), "CITYWIDE", citywide_grid)

    # ── District sheets ──────────────────────────────────────────────────────────
    for dist in ['North', 'East', 'South', 'West']:
        dist_grid = build_district_grid(category_counts, dist, month_key, week_num=args.week)
        write_sheet(wb.create_sheet(dist.upper()), dist.upper(), dist_grid)

    # ── Shots Fired & Shootings sheet ────────────────────────────────────────────
    shots_grid = build_shots_sheet(all_incidents, category_counts)
    write_sheet(wb.create_sheet("Shots Fired & Shootings"), "Shots Fired & Shootings", shots_grid)

    # ── MVA Data sheet ───────────────────────────────────────────────────────────
    mva_grid = build_mva_sheet(all_incidents)
    write_sheet(wb.create_sheet("MVA Data"), "MVA Data", mva_grid)

    # ── Save ─────────────────────────────────────────────────────────────────────
    wb.save(args.output)
    print(f"\n✓  Workbook written: {args.output}")

# ─── Grid builders ──────────────────────────────────────────────────────────────

def build_citywide_grid(counts, month_key, week_num):
    rows = []
    # Header rows (same as template)
    rows.append(['JERSEY CITY POLICE DEPARTMENT — JC IMPACT   |   WEEKLY CRIME BREAKDOWN   |   CITYWIDE', '', '', '', '', '', '', '', '', ''])
    rows.append([f'Source: JCPD CompStat   |   Data is preliminary / subject to change   |   Prepared by: Geremy Munoz   |   Last Updated: {date.today().strftime("%m/%d/%Y")}', '', '', '', '', '', '', '', '', ''])
    rows.append(['', '', '', '', '', '', '', '', '', ''])
    rows.append([f'▌  {month_label(month_key)}   ({week_range_str(month_key, week_num)})', '', '', '', '', '', '', '', '', ''])
    rows.append(['OFFENSE', 'Note', 'Week 1', 'Week 2', 'Week 3', 'Week 4', 'Week 5', 'MON Total', '', ''])
    rows.append(['★ 7 MAJOR IMPACT CATEGORIES', '', '', '', '', '', '', '', '', ''])
    # For each major category
    for cat in ['assault total', 'burglary total', 'theft', 'robbery total', 'stolen vehicles', 'homicide', 'sex offenses']:
        row = [cat.title()]
        # week columns blank + total
        row += ['']*5 + [counts.get(cat, {}).get('citywide', 0)] + ['']*2
        rows.append(row)
    # Separator
    rows.append(['', '', '', '', '', '', '', '', '', ''])
    # Supplementals
    for sup in ['gun recoveries', 'shoplifting over 500', 'shots fired', 'mva']:
        rows.append([sup.title(), '', '', '', '', '', '', counts.get(sup,{}).get('citywide',0), '', ''])
    return rows

def build_district_grid(counts, district, month_key, week_num):
    rows = []
    rows.append([f'JERSEY CITY POLICE DEPARTMENT — JC IMPACT   |   WEEKLY CRIME BREAKDOWN   |   {district.upper()}', '', '', '', '', '', '', '', '', ''])
    rows.append([f'Source: JCPD CompStat   |   Data is preliminary / subject to change   |   Prepared by: Geremy Munoz   |   Last Updated: {date.today().strftime("%m/%d/%Y")}', '', '', '', '', '', '', '', '', ''])
    rows.append(['', '', '', '', '', '', '', '', '', ''])
    rows.append([f'▌  {month_label(month_key)}   ({week_range_str(month_key, week_num)})', '', '', '', '', '', '', '', '', ''])
    rows.append(['OFFENSE', 'Note', 'Week 1', 'Week 2', 'Week 3', 'Week 4', 'Week 5', 'MON Total', '', ''])
    rows.append(['★ 7 MAJOR IMPACT CATEGORIES', '', '', '', '', '', '', '', '', ''])
    for cat in ['assault total', 'burglary total', 'theft', 'robbery total', 'stolen vehicles', 'homicide', 'sex offenses']:
        dist_cnt = counts.get(cat, {}).get('districts', {}).get(district, 0)
        row = [cat.title(), '', '', '', '', '', '', dist_cnt, '', '']
        rows.append(row)
    return rows

def build_shots_sheet(incidents, counts):
    rows = []
    rows.append(['SHOTS FIRED & SHOOTINGS', '', '', '', '', '', '', '', '', ''])
    rows.append(['Date', 'Case #', 'Location', 'District', 'Description', 'Lat', 'Lng', '', '', ''])
    for inc in incidents:
        if inc.get('category') in ('shots_fired', 'shooting'):
            row = [
                inc.get('date', ''),
                inc.get('case', ''),
                inc.get('address', ''),
                inc.get('district', ''),
                inc.get('description', ''),
                inc.get('lat', ''),
                inc.get('lng', ''),
            ]
            # pad to 10 cols
            row += [''] * (10 - len(row))
            rows.append(row)
    return rows

def build_mva_sheet(incidents):
    rows = []
    rows.append(['MVA Data', '', '', '', '', '', '', '', '', ''])
    rows.append(['Case #', 'Date', 'Road', 'Cross', 'District', 'Lat', 'Lng', '', '', ''])
    for inc in incidents:
        if inc.get('category') == 'mva':
            row = [
                inc.get('case', ''),
                inc.get('date', ''),
                inc.get('road', ''),
                inc.get('cross', ''),
                inc.get('district', ''),
                inc.get('lat', ''),
                inc.get('lng', ''),
            ]
            row += [''] * (10 - len(row))
            rows.append(row)
    return rows

# ─── Helpers ────────────────────────────────────────────────────────────────────
def find_file(folder, pattern, flags=0):
    import re
    pat = re.compile(pattern, flags)
    for f in Path(folder).iterdir():
        if pat.search(f.name):
            return str(f)
    return None

def month_label(month_key):
    import calendar
    year, month = month_key.split('-')
    month_name = calendar.month_abbr[int(month)]
    return f"{month_name} '{year[2:]}"

def week_range_str(month_key, week_num):
    # Hardcode for now; in reality would compute from week number
    # For Week 16 (Apr 12-18 2026)
    weeks_in_month = {
        '2026-04': {1: 'Apr 1-3', 2: 'Apr 4-10', 3: 'Apr 11-17', 4: 'Apr 18-24', 5: 'Apr 25-30'},
        '2026-03': {1: 'Mar 1-3', 2: 'Mar 4-10', 3: 'Mar 11-17', 4: 'Mar 18-24', 5: 'Mar 25-31'},
    }
    return weeks_in_month.get(month_key, {}).get(week_num, f'Week {week_num}')

if __name__ == '__main__':
    main()
